"""Phase 6.3 / 6.4: Excel export for IPMT and WAR using openpyxl."""
import io
import os
import logging
from datetime import date

from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from .war_config import get_war_table_config

logger = logging.getLogger(__name__)


def _month_range(year: int, month: int):
    """First and last day of the given month."""
    start = date(year, month, 1)
    if month == 12:
        end = date(year, 12, 31)
    else:
        end = date(year, month + 1, 1)
    from datetime import timedelta
    end = end - timedelta(days=1)
    return start, end


def build_ipmt_excel(personnel, year: int, month: int, preview_rows=None):
    """
    Build IPMT Excel: WARs for the given personnel where period overlaps the given month.
    Columns: Request ID, Period Start, Period End, Summary, Accomplishments, Success Indicators.
    """
    from .models import WorkAccomplishmentReport
    start, end = _month_range(year, month)
    # WARs where personnel matches and (period_start <= end and period_end >= start)
    qs = WorkAccomplishmentReport.objects.filter(
        personnel=personnel,
        period_start__lte=end,
        period_end__gte=start,
    ).select_related('request', 'personnel').prefetch_related('success_indicators').order_by('period_start')
    return _ipmt_list_to_excel(
        personnel=personnel,
        queryset=qs,
        title=f"IPMT Report — {personnel.get_full_name() or personnel.username} — {year}-{month:02d}",
        sheet_name=f"IPMT {year}-{month:02d}",
        year=year,
        month=month,
        rows_override=preview_rows or [],
    )


def _month_label(year: int, month: int):
    start, end = _month_range(year, month)
    return f"{start.strftime('%B').upper()} {start.day}-{end.day}, {year}"


def _ipmt_indicator_rows(queryset):
    """
    Build rows for IPMT table:
    [
      (indicator_text, [accomplishment1, accomplishment2, ...], comment_text),
    ]
    """
    rows = []
    grouped = {}
    for war in queryset:
        indicators = list(war.success_indicators.all())
        if not indicators:
            continue
        accomplishment_text = (war.accomplishments or war.summary or '').strip()
        if not accomplishment_text:
            accomplishment_text = (
                f"{war.period_start:%b %d} - {war.period_end:%b %d, %Y}: "
                f"{getattr(getattr(war, 'request', None), 'title', '') or 'Completed work activity'}"
            )
        for indicator in indicators:
            indicator_key = indicator.pk
            if indicator_key not in grouped:
                label = f"{indicator.code}. {indicator.name}" if indicator.code else indicator.name
                grouped[indicator_key] = {
                    "indicator": label,
                    "order": indicator.display_order or 0,
                    "accomplishments": [],
                }
            grouped[indicator_key]["accomplishments"].append(accomplishment_text)

    if not grouped:
        return [("No success indicators tagged yet.", ["Tag indicators in WAR first, then regenerate IPMT."], "")]

    for _, payload in sorted(grouped.items(), key=lambda item: (item[1]["order"], item[1]["indicator"].lower())):
        unique_accomplishments = []
        seen = set()
        for item in payload["accomplishments"]:
            key = item.lower()
            if key in seen:
                continue
            seen.add(key)
            unique_accomplishments.append(item)
        rows.append((payload["indicator"], unique_accomplishments, "Complied"))
    return rows


def _ipmt_list_to_excel(personnel, queryset, title="IPMT", sheet_name="IPMT", year=None, month=None, rows_override=None):
    """
    Build IPMT workbook matching the provided sample layout:
    metadata rows + 3-column indicator table + signature block.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "IPMT")[:31]

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    wrap_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 9-column layout to mirror the formal IPMT header sample.
    # A-C: Success Indicators | D-G: Accomplishments | H-I: Comments/Remarks
    widths = {
        "A": 18, "B": 18, "C": 18,
        "D": 18, "E": 18, "F": 18, "G": 18,
        "H": 14, "I": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Header metadata block (employee + period values)
    month_text = _month_label(year, month) if year and month else _format_period_range(queryset)
    unit_name = (
        getattr(getattr(personnel, "unit", None), "name", None)
        or "General Services Office"
    )
    employee_name = personnel.get_full_name() or getattr(personnel, "username", "") or "—"
    employment_status = (getattr(personnel, "employment_status", "") or "—").upper()
    position_title = (getattr(personnel, "position_title", "") or "—").upper()

    # ----- Formal top header -----
    ws.merge_cells("A1:A4")
    ws.merge_cells("B1:F2")
    ws.merge_cells("B3:F4")

    ws["B1"] = "INDIVIDUAL PERFORMANCE\nMONITORING TOOLS"
    ws["B1"].font = Font(name="Calibri", size=20, bold=True)
    ws["B1"].alignment = wrap_center

    ws["B3"] = "Work Accomplishment Report"
    ws["B3"].font = Font(name="Calibri", size=14)
    ws["B3"].alignment = wrap_center

    # Right metadata box
    right_labels = [
        ("G1:H1", "Doc. Ref. No.:", "I1", ""),
        ("G2:H2", "Effective Date:", "I2", "June 30, 2023"),
        ("G3:H3", "Revision No.:", "I3", "00"),
        ("G4:H4", "Page No.:", "I4", "Page 1 of ___"),
    ]
    for label_merge, label_text, value_cell, value_text in right_labels:
        ws.merge_cells(label_merge)
        ws[label_merge.split(":")[0]] = label_text
        ws[label_merge.split(":")[0]].font = Font(name="Calibri", size=10)
        ws[label_merge.split(":")[0]].alignment = Alignment(horizontal="left", vertical="center")
        ws[value_cell] = value_text
        ws[value_cell].font = Font(name="Calibri", size=10, bold=True)
        ws[value_cell].alignment = wrap_center

    # Left logo if available
    psu_path, _ = _resolve_logo_paths()
    try:
        from openpyxl.drawing.image import Image
        if psu_path and os.path.isfile(psu_path):
            img_psu = Image(psu_path)
            # Keep logo visually centered inside the merged A1:A4 header box.
            img_psu.height = 116
            img_psu.width = 116
            ws.add_image(img_psu, "A1")
    except (ImportError, OSError, ValueError):
        logger.exception('Unable to load IPMT logo image assets.')

    # ----- Employee information rows -----
    info_rows = [
        (5, "College/Campus/Department/Unit :", unit_name.upper()),
        (6, "Name of Employee :", employee_name.upper()),
        (7, "Status of Employment :", employment_status),
        (8, "Position :", position_title),
        (9, "Month:", (month_text or "—").upper()),
    ]
    for r, label, value in info_rows:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
        ws.cell(row=r, column=1, value=label).font = Font(name="Calibri", size=12, bold=True)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=r, column=4, value=value).font = Font(name="Calibri", size=12)
        ws.cell(row=r, column=4).alignment = wrap_left

    # ----- Table header row -----
    table_header_row = 10
    ws.merge_cells(start_row=table_header_row, start_column=1, end_row=table_header_row, end_column=3)
    ws.merge_cells(start_row=table_header_row, start_column=4, end_row=table_header_row, end_column=7)
    ws.merge_cells(start_row=table_header_row, start_column=8, end_row=table_header_row, end_column=9)
    ws.cell(row=table_header_row, column=1, value="*Success Indicators\n(Based on the IPCR Targets)")
    ws.cell(row=table_header_row, column=4, value="Actual Accomplishments")
    ws.cell(row=table_header_row, column=8, value="Comments / Remarks")
    for c in (1, 4, 8):
        cell = ws.cell(row=table_header_row, column=c)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.alignment = wrap_center
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    row = table_header_row + 1
    source_rows = []
    if rows_override:
        for item in rows_override:
            indicator_text = str(item.get("indicator", "")).strip()
            if not indicator_text:
                continue
            accomplishments = item.get("accomplishments") or [""]
            if not isinstance(accomplishments, list):
                accomplishments = [str(accomplishments)]
            clean_accomplishments = [str(acc) for acc in accomplishments]
            source_rows.append((indicator_text, clean_accomplishments, str(item.get("comment", "")).strip()))
    else:
        source_rows = _ipmt_indicator_rows(queryset)

    for indicator_text, accomplishments, comment in source_rows:
        start_row = row
        entries = accomplishments or [""]
        for idx, accomplishment in enumerate(entries):
            if idx == 0:
                ws.cell(row=row, column=1, value=indicator_text)
            else:
                ws.cell(row=row, column=1, value=None)
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)
            ws.cell(row=row, column=4, value=accomplishment)
            ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
            ws.cell(row=row, column=8, value=comment if idx == 0 else "")
            row += 1
        if row - start_row > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=row - 1, end_column=3)
        else:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
        for r in range(start_row, row):
            for c in range(1, 10):
                cell = ws.cell(row=r, column=c)
                cell.font = Font(name="Calibri", size=10)
                if c <= 7:
                    cell.alignment = wrap_left
                else:
                    cell.alignment = wrap_center
                cell.border = thin_border

    # Note row
    note_row = row + 1
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=9)
    ws.cell(
        row=note_row,
        column=1,
        value="*Based on the IPCR Major Final Output (MFO)/ Program, Activity and Project (PAP), "
              "select only those success indicators where the accomplishments for the period are aligned to.",
    )
    note_cell = ws.cell(row=note_row, column=1)
    note_cell.font = Font(name="Calibri", size=9, italic=True)
    note_cell.alignment = wrap_left
    note_cell.border = thin_border
    for c in range(2, 10):
        ws.cell(row=note_row, column=c).border = thin_border

    # Signature block
    sig_title_row = note_row + 2
    ws.merge_cells(start_row=sig_title_row, start_column=1, end_row=sig_title_row, end_column=3)
    ws.merge_cells(start_row=sig_title_row, start_column=7, end_row=sig_title_row, end_column=9)
    ws.cell(row=sig_title_row, column=1, value="Prepared by:").font = Font(name="Calibri", size=10, bold=True)
    ws.cell(row=sig_title_row, column=7, value="Checked and Verified by:").font = Font(name="Calibri", size=10, bold=True)

    sig_name_row = sig_title_row + 3
    ws.merge_cells(start_row=sig_name_row, start_column=1, end_row=sig_name_row, end_column=3)
    ws.merge_cells(start_row=sig_name_row, start_column=7, end_row=sig_name_row, end_column=9)
    ws.cell(row=sig_name_row, column=1, value=employee_name).font = Font(name="Calibri", size=10, bold=True)

    from apps.gso_accounts.models import User
    supervisor = User.objects.filter(role=User.Role.DIRECTOR, is_active=True).order_by('id').first()
    supervisor_name = (supervisor.get_full_name() if supervisor else "") or ""
    ws.cell(row=sig_name_row, column=7, value=supervisor_name.upper()).font = Font(name="Calibri", size=10, bold=True)

    sig_role_row = sig_name_row + 1
    ws.merge_cells(start_row=sig_role_row, start_column=1, end_row=sig_role_row, end_column=3)
    ws.merge_cells(start_row=sig_role_row, start_column=7, end_row=sig_role_row, end_column=9)
    ws.cell(row=sig_role_row, column=1, value="Employee").font = Font(name="Calibri", size=10)
    ws.cell(row=sig_role_row, column=7, value="Department Head/ Supervisor").font = Font(name="Calibri", size=10)

    # Border + row height tuning
    for r in range(1, sig_role_row + 1):
        if r >= table_header_row:
            ws.row_dimensions[r].height = 34
        else:
            ws.row_dimensions[r].height = 22
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[note_row].height = 36

    # Add borders for entire printable header and metadata area.
    for r in range(1, table_header_row):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            if r >= 5:
                cell.alignment = wrap_left if c <= 7 else wrap_center

    # Make the title + subtitle feel like one connected block:
    # keep only the outer border for B1:F4, remove internal divider lines.
    border_none = Side(style=None)
    for r in range(1, 5):
        for c in range(2, 7):
            ws.cell(row=r, column=c).border = Border(
                left=thin_border.left if c == 2 else border_none,
                right=thin_border.right if c == 6 else border_none,
                top=thin_border.top if r == 1 else border_none,
                bottom=thin_border.bottom if r == 4 else border_none,
            )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, title.replace(" ", "_")[:80]


def build_war_export_excel(queryset, title="WAR Export", unit=None, split_by_unit_when_all=False):
    """Build WAR export Excel. If all units is selected, optionally create one sheet per unit."""
    if unit is None and split_by_unit_when_all:
        return _war_list_to_excel_by_unit(queryset, title=title)
    _config_key, config = get_war_table_config(unit)
    return _war_list_to_excel(
        queryset,
        title=title,
        sheet_name="WAR",
        unit=unit,
        table_config=config,
    )


def _format_period_range(queryset):
    """Return a string like 'AUGUST 1-31, 2025' from queryset min/max period dates."""
    if not queryset:
        return ""
    starts = [w.period_start for w in queryset]
    ends = [w.period_end for w in queryset]
    d_start = min(starts)
    d_end = max(ends)
    if d_start.month == d_end.month and d_start.year == d_end.year:
        return f"{d_start.strftime('%B').upper()} {d_start.day}-{d_end.day}, {d_start.year}"
    return f"{d_start.strftime('%B %d, %Y').upper()} – {d_end.strftime('%B %d, %Y').upper()}"


def _resolve_logo_paths():
    """Return absolute paths for PSU and GSO logos under static/img/logo."""
    base = getattr(settings, "BASE_DIR", None)
    if base is None:
        return None, None
    try:
        psu_path = base / "static" / "img" / "logo" / "psu_logo.png"
        gso_path = base / "static" / "img" / "logo" / "gso_logo.png"
    except TypeError:
        psu_path = os.path.join(base, "static", "img", "logo", "psu_logo.png")
        gso_path = os.path.join(base, "static", "img", "logo", "gso_logo.png")
    return str(psu_path), str(gso_path)


def _get_requesting_office_name(req):
    """Resolve requesting office/department from requestor profile."""
    requestor = getattr(req, "requestor", None)
    if not requestor:
        return ""
    office = (getattr(requestor, "office_department", "") or "").strip()
    if office:
        return office
    # Export should represent office source, not personal identity.
    return ""


def _apply_standard_header(ws, total_columns, period_label="", report_label="WORK ACCOMPLISHMENT REPORT", unit_label="ALL UNITS"):
    """
    Apply the common PSU-GSO Excel header used by generated exports.
    Returns table header row index (int).
    """
    max_col = max(1, total_columns)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in range(1, 11):
        ws.row_dimensions[r].height = 18
    ws.row_dimensions[5].height = 13.5
    ws.row_dimensions[6].height = 17.4
    ws.row_dimensions[9].height = 17.4
    ws.row_dimensions[10].height = 31.5

    # Institutional text block
    institutional_lines = [
        ("Republic of the Philippines", Font(name="Arial Narrow", size=14)),
        ("PALAWAN STATE UNIVERSITY", Font(name="Arial Narrow", size=14, bold=True)),
        ("GENERAL SERVICES OFFICE", Font(name="Arial Narrow", size=14, bold=True)),
        ("Puerto Princesa City", Font(name="Arial Narrow", size=14)),
    ]
    for row, (text, font) in enumerate(institutional_lines, start=1):
        c = ws.cell(row=row, column=1, value=text)
        c.font = font
        c.alignment = center
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)

    # Report title block
    report = ws.cell(row=6, column=1, value=report_label)
    report.font = Font(name="Broadway", size=14)
    report.alignment = center
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=max_col)

    period = ws.cell(row=7, column=1, value=period_label or "—")
    period.font = Font(name="Calibri", size=14, bold=True)
    period.alignment = center
    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=max_col)

    unit = ws.cell(row=9, column=1, value=unit_label)
    unit.font = Font(name="Arial", size=10, bold=True)
    unit.alignment = center
    ws.merge_cells(start_row=9, start_column=1, end_row=9, end_column=max_col)

    # Add logos if files are present.
    psu_path, gso_path = _resolve_logo_paths()
    try:
        from openpyxl.drawing.image import Image
        logo_height_px = 95
        # Choose logo anchors by sheet width so headers stay clean on wider reports
        # (e.g., Feedback export has many columns).
        if max_col >= 14:
            # Wide sheets (e.g., Feedback): keep logos near centered header text,
            # matching the approved sample layout.
            center_col = max(1, (max_col + 1) // 2)
            left_logo_anchor = f"{get_column_letter(max(1, center_col - 3))}1"
            right_logo_anchor = f"{get_column_letter(min(max_col, center_col + 3))}1"
        else:
            # WAR sheets: keep approved positioning close to center text block.
            left_logo_anchor = "D1"
            right_logo_anchor = "G1"
        if psu_path and os.path.isfile(psu_path):
            img_psu = Image(psu_path)
            w_orig, h_orig = img_psu.width, img_psu.height
            img_psu.height = logo_height_px
            img_psu.width = int(w_orig * (logo_height_px / h_orig))
            if max_col >= 10:
                ws.add_image(img_psu, left_logo_anchor)
            else:
                ws.add_image(img_psu, "A1")
        if gso_path and os.path.isfile(gso_path):
            img_gso = Image(gso_path)
            w_orig, h_orig = img_gso.width, img_gso.height
            img_gso.height = logo_height_px
            img_gso.width = int(w_orig * (logo_height_px / h_orig))
            if max_col >= 10:
                ws.add_image(img_gso, right_logo_anchor)
            else:
                right_col = get_column_letter(max(1, max_col - 1))
                ws.add_image(img_gso, f"{right_col}1")
    except (ImportError, OSError, ValueError):
        # Export should remain functional even when image loading is unavailable.
        logger.exception('Unable to load WAR/feedback header logo image assets.')

    return 10


def _war_list_to_excel(queryset, title="WAR", sheet_name="WAR", unit=None, table_config=None):
    """
    Write WAR queryset to an openpyxl workbook. Uses table_config for headers and column count;
    if None, uses generic config from get_war_table_config(unit).
    Returns (buf, filename_suffix).
    """
    if table_config is None:
        _, table_config = get_war_table_config(unit)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    period_label = _format_period_range(queryset) if queryset else ""
    headers = list(table_config['excel_headers'])
    num_cols = table_config['excel_column_count']
    layout_label = (table_config.get('label') or '').strip().lower() if table_config else ''
    if layout_label == 'repair and maintenance' and num_cols == 11:
        headers = [
            'Date Started',
            'Date Completed',
            'Name of Project',
            'Description',
            'Requesting Office',
            'Assigned Personnel',
            'Status',
            'Material Cost',
            'Labor Cost',
            'Total Cost',
            'Control #',
        ]
    if len(headers) < num_cols:
        headers.extend([''] * (num_cols - len(headers)))
    elif len(headers) > num_cols:
        headers = headers[:num_cols]
    table_header_row = _apply_standard_header(
        ws,
        total_columns=num_cols,
        period_label=period_label,
        report_label="WORK ACCOMPLISHMENT REPORT",
        unit_label=(unit.name.upper() if unit else "ALL UNITS"),
    )

    # Row after common header: table headers (from table_config)
    header_fill = table_config.get('header_fill')
    fill = PatternFill(start_color=header_fill, end_color=header_fill, fill_type='solid') if header_fill else None
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=table_header_row, column=col, value=h)
        cell.font = Font(name="Arial Narrow", size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        if fill:
            cell.fill = fill
    data_start_row = table_header_row + 1
    row = data_start_row
    for war in queryset:
        req = war.request
        req_display_id = getattr(req, "display_id", None)
        if callable(req_display_id):
            request_id_value = req_display_id()
        elif req_display_id:
            request_id_value = req_display_id
        else:
            request_id_value = str(req.pk)
        requesting_office = _get_requesting_office_name(req)
        total_materials = war.material_cost
        labor = war.labor_cost
        total = war.total_cost
        if num_cols == 12:
            # Generic (All units): Request ID, Unit, Personnel, Date Started, Date Completed, Name of Activity, Description, Requesting Office, Status, Total Materials, Labor, Total
            ws.cell(row=row, column=1, value=request_id_value)
            ws.cell(row=row, column=2, value=req.unit.name if req.unit_id else "")
            ws.cell(row=row, column=3, value=war.personnel.get_full_name() or getattr(war.personnel, 'username', ''))
            ws.cell(row=row, column=4, value=war.period_start)
            ws.cell(row=row, column=5, value=war.period_end)
            ws.cell(row=row, column=6, value=war.summary or "")
            ws.cell(row=row, column=7, value=war.accomplishments or "")
            ws.cell(row=row, column=8, value=requesting_office)
            ws.cell(row=row, column=9, value="Completed")
            ws.cell(row=row, column=10, value=total_materials)
            if total_materials is not None:
                ws.cell(row=row, column=10).number_format = "#,##0.00"
            ws.cell(row=row, column=11, value=labor)
            if labor is not None:
                ws.cell(row=row, column=11).number_format = "#,##0.00"
            ws.cell(row=row, column=12, value=total)
            if total is not None:
                ws.cell(row=row, column=12).number_format = "#,##0.00"
        elif num_cols == 11 and layout_label == 'repair and maintenance':
            # Repair & Maintenance (11 columns): includes Control # in K.
            personnel_name = war.personnel.get_full_name() or getattr(war.personnel, 'username', '')
            ws.cell(row=row, column=1, value=war.period_start)
            ws.cell(row=row, column=2, value=war.period_end)
            ws.cell(row=row, column=3, value=war.summary or "")
            ws.cell(row=row, column=4, value=war.accomplishments or "")
            ws.cell(row=row, column=5, value=requesting_office)
            ws.cell(row=row, column=6, value=personnel_name)
            ws.cell(row=row, column=7, value="Completed")
            ws.cell(row=row, column=8, value=total_materials)
            if total_materials is not None:
                ws.cell(row=row, column=8).number_format = "#,##0.00"
            ws.cell(row=row, column=9, value=labor)
            if labor is not None:
                ws.cell(row=row, column=9).number_format = "#,##0.00"
            ws.cell(row=row, column=10, value=total)
            if total is not None:
                ws.cell(row=row, column=10).number_format = "#,##0.00"
            ws.cell(row=row, column=11, value=request_id_value)
        elif num_cols == 11:
            # Electrical (11 columns): Date Started, Date Complete, Name of Project, Description,
            # Requesting Office, Assigned Personnel, Status, Material Cost, Labor Cost, Total Cost, Control #
            personnel_name = war.personnel.get_full_name() or getattr(war.personnel, 'username', '')
            ws.cell(row=row, column=1, value=war.period_start)
            ws.cell(row=row, column=2, value=war.period_end)
            ws.cell(row=row, column=3, value=war.summary or "")
            ws.cell(row=row, column=4, value=war.accomplishments or "")
            ws.cell(row=row, column=5, value=requesting_office)
            ws.cell(row=row, column=6, value=personnel_name)
            ws.cell(row=row, column=7, value="Done")
            ws.cell(row=row, column=8, value=total_materials)
            if total_materials is not None:
                ws.cell(row=row, column=8).number_format = "#,##0.00"
            ws.cell(row=row, column=9, value=labor)
            if labor is not None:
                ws.cell(row=row, column=9).number_format = "#,##0.00"
            ws.cell(row=row, column=10, value=total)
            if total is not None:
                ws.cell(row=row, column=10).number_format = "#,##0.00"
            ws.cell(row=row, column=11, value=request_id_value)
        else:
            # Fallback layout if a new config is added later.
            ws.cell(row=row, column=1, value=request_id_value)
            ws.cell(row=row, column=2, value=war.period_start)
            ws.cell(row=row, column=3, value=war.period_end)
            ws.cell(row=row, column=4, value=war.summary or "")
            ws.cell(row=row, column=5, value=war.accomplishments or "")
        for c in range(1, num_cols + 1):
            data_cell = ws.cell(row=row, column=c)
            data_cell.border = thin_border
            data_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        row += 1
    # Column widths by layout
    if num_cols == 12:
        widths = [12, 18, 18, 14, 14, 22, 35, 18, 12, 14, 12, 18]
    elif num_cols == 11 and layout_label == 'repair and maintenance':
        widths = [14, 14, 22, 35, 18, 18, 12, 14, 12, 18, 11.6]
    elif num_cols == 11:
        # Closer to the provided sample WAR electrical sheet.
        widths = [7.9, 13, 11.7, 23.1, 14, 12.9, 8, 8.3, 13, 13, 11.6]
    else:
        widths = [14, 14, 20, 35, 18]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    # Recompute row heights after final column widths are known.
    for data_row in range(data_start_row, row):
        _adjust_war_row_height(ws, row=data_row, num_cols=num_cols, layout_label=layout_label)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, title.replace(" ", "_")[:80]


def _war_list_to_excel_by_unit(queryset, title="WAR"):
    """Create a workbook with one sheet per unit from the provided queryset."""
    war_rows = list(queryset)
    units_by_id = {}
    for war in war_rows:
        req_unit = getattr(getattr(war, 'request', None), 'unit', None)
        if req_unit and req_unit.id not in units_by_id:
            units_by_id[req_unit.id] = req_unit

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    if not units_by_id:
        ws = wb.create_sheet(title="WAR")
        _config_key, config = get_war_table_config(None)
        _write_war_rows_to_sheet(ws, [], unit=None, table_config=config)
    else:
        sorted_units = sorted(units_by_id.values(), key=lambda u: (u.name or "").lower())
        for unit in sorted_units:
            unit_rows = [w for w in war_rows if getattr(getattr(w, 'request', None), 'unit_id', None) == unit.id]
            sheet_name = (unit.code or unit.name or "Unit")[:31]
            ws = wb.create_sheet(title=sheet_name)
            _config_key, config = get_war_table_config(unit)
            _write_war_rows_to_sheet(ws, unit_rows, unit=unit, table_config=config)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, title.replace(" ", "_")[:80]


def _write_war_rows_to_sheet(ws, war_rows, unit=None, table_config=None):
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    period_label = _format_period_range(war_rows) if war_rows else ""
    headers = list(table_config['excel_headers'])
    num_cols = table_config['excel_column_count']
    layout_label = (table_config.get('label') or '').strip().lower() if table_config else ''
    if layout_label == 'repair and maintenance' and num_cols == 11:
        headers = [
            'Date Started',
            'Date Completed',
            'Name of Project',
            'Description',
            'Requesting Office',
            'Assigned Personnel',
            'Status',
            'Material Cost',
            'Labor Cost',
            'Total Cost',
            'Control #',
        ]
    if len(headers) < num_cols:
        headers.extend([''] * (num_cols - len(headers)))
    elif len(headers) > num_cols:
        headers = headers[:num_cols]
    table_header_row = _apply_standard_header(
        ws,
        total_columns=num_cols,
        period_label=period_label,
        report_label="WORK ACCOMPLISHMENT REPORT",
        unit_label=(unit.name.upper() if unit else "ALL UNITS"),
    )
    header_fill = table_config.get('header_fill')
    fill = PatternFill(start_color=header_fill, end_color=header_fill, fill_type='solid') if header_fill else None
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=table_header_row, column=col, value=h)
        cell.font = Font(name="Arial Narrow", size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        if fill:
            cell.fill = fill
    row = table_header_row + 1
    for war in war_rows:
        req = war.request
        req_display_id = getattr(req, "display_id", None)
        if callable(req_display_id):
            request_id_value = req_display_id()
        elif req_display_id:
            request_id_value = req_display_id
        else:
            request_id_value = str(req.pk)
        requesting_office = _get_requesting_office_name(req)
        total_materials = war.material_cost
        labor = war.labor_cost
        total = war.total_cost
        if num_cols == 12:
            ws.cell(row=row, column=1, value=request_id_value)
            ws.cell(row=row, column=2, value=req.unit.name if req.unit_id else "")
            ws.cell(row=row, column=3, value=war.personnel.get_full_name() or getattr(war.personnel, 'username', ''))
            ws.cell(row=row, column=4, value=war.period_start)
            ws.cell(row=row, column=5, value=war.period_end)
            ws.cell(row=row, column=6, value=war.summary or "")
            ws.cell(row=row, column=7, value=war.accomplishments or "")
            ws.cell(row=row, column=8, value=requesting_office)
            ws.cell(row=row, column=9, value="Completed")
            ws.cell(row=row, column=10, value=total_materials)
            if total_materials is not None:
                ws.cell(row=row, column=10).number_format = "#,##0.00"
            ws.cell(row=row, column=11, value=labor)
            if labor is not None:
                ws.cell(row=row, column=11).number_format = "#,##0.00"
            ws.cell(row=row, column=12, value=total)
            if total is not None:
                ws.cell(row=row, column=12).number_format = "#,##0.00"
        elif num_cols == 11 and layout_label == 'repair and maintenance':
            ws.cell(row=row, column=1, value=war.period_start)
            ws.cell(row=row, column=2, value=war.period_end)
            ws.cell(row=row, column=3, value=war.summary or "")
            ws.cell(row=row, column=4, value=war.accomplishments or "")
            ws.cell(row=row, column=5, value=requesting_office)
            ws.cell(row=row, column=6, value=war.personnel.get_full_name() or getattr(war.personnel, 'username', ''))
            ws.cell(row=row, column=7, value="Completed")
            ws.cell(row=row, column=8, value=total_materials)
            if total_materials is not None:
                ws.cell(row=row, column=8).number_format = "#,##0.00"
            ws.cell(row=row, column=9, value=labor)
            if labor is not None:
                ws.cell(row=row, column=9).number_format = "#,##0.00"
            ws.cell(row=row, column=10, value=total)
            if total is not None:
                ws.cell(row=row, column=10).number_format = "#,##0.00"
            ws.cell(row=row, column=11, value=request_id_value)
        elif num_cols == 11:
            personnel_name = war.personnel.get_full_name() or getattr(war.personnel, 'username', '')
            ws.cell(row=row, column=1, value=war.period_start)
            ws.cell(row=row, column=2, value=war.period_end)
            ws.cell(row=row, column=3, value=war.summary or "")
            ws.cell(row=row, column=4, value=war.accomplishments or "")
            ws.cell(row=row, column=5, value=requesting_office)
            ws.cell(row=row, column=6, value=personnel_name)
            ws.cell(row=row, column=7, value="Done")
            ws.cell(row=row, column=8, value=total_materials)
            if total_materials is not None:
                ws.cell(row=row, column=8).number_format = "#,##0.00"
            ws.cell(row=row, column=9, value=labor)
            if labor is not None:
                ws.cell(row=row, column=9).number_format = "#,##0.00"
            ws.cell(row=row, column=10, value=total)
            if total is not None:
                ws.cell(row=row, column=10).number_format = "#,##0.00"
            ws.cell(row=row, column=11, value=request_id_value)
        else:
            ws.cell(row=row, column=1, value=request_id_value)
            ws.cell(row=row, column=2, value=war.period_start)
            ws.cell(row=row, column=3, value=war.period_end)
            ws.cell(row=row, column=4, value=war.summary or "")
            ws.cell(row=row, column=5, value=war.accomplishments or "")
        for c in range(1, num_cols + 1):
            data_cell = ws.cell(row=row, column=c)
            data_cell.border = thin_border
            data_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        row += 1
    if num_cols == 12:
        widths = [12, 18, 18, 14, 14, 22, 35, 18, 12, 14, 12, 18]
    elif num_cols == 11 and layout_label == 'repair and maintenance':
        widths = [14, 14, 22, 35, 18, 18, 12, 14, 12, 18, 11.6]
    elif num_cols == 11:
        widths = [7.9, 13, 11.7, 23.1, 14, 12.9, 8, 8.3, 13, 13, 11.6]
    else:
        widths = [14, 14, 20, 35, 18]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    for data_row in range(table_header_row + 1, row):
        _adjust_war_row_height(ws, row=data_row, num_cols=num_cols, layout_label=layout_label)


def _adjust_war_row_height(ws, row, num_cols, layout_label):
    """Increase row height based on description cell length/newlines."""
    # Description column by layout:
    # generic(12): col 7, repair/electrical(11): col 4
    if num_cols == 12:
        desc_col = 7
        chars_per_line = 55
    elif num_cols == 11 and layout_label in ('repair and maintenance', 'electrical services unit'):
        desc_col = 4
        chars_per_line = 38
    else:
        return
    cell = ws.cell(row=row, column=desc_col)
    text = cell.value or ""
    if not isinstance(text, str):
        text = str(text)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if not text:
        return
    # Use effective column width for more accurate wrapping.
    # Excel width units are roughly character-count based for Calibri 11.
    col_letter = get_column_letter(desc_col)
    col_width = ws.column_dimensions[col_letter].width or chars_per_line
    usable_chars = max(10, int(col_width * 1.15))

    import textwrap
    wrapped_lines = 0
    for paragraph in text.splitlines() or [""]:
        if not paragraph:
            wrapped_lines += 1
            continue
        wrapped_lines += max(1, len(textwrap.wrap(paragraph, width=usable_chars, break_long_words=True, break_on_hyphens=False)))

    # 15 points is default one-line height. Add slight padding.
    ws.row_dimensions[row].height = max(20, min(409, wrapped_lines * 15 + 4))


def get_war_queryset(personnel=None, unit=None, date_from=None, date_to=None):
    """Return WorkAccomplishmentReport queryset filtered by optional personnel, unit, and date range."""
    from .models import WorkAccomplishmentReport
    qs = WorkAccomplishmentReport.objects.select_related(
        'request', 'personnel', 'request__unit', 'request__requestor',
    ).prefetch_related('success_indicators').order_by('-period_end', '-created_at')
    if personnel:
        qs = qs.filter(personnel=personnel)
    if unit:
        qs = qs.filter(request__unit=unit)
    if date_from:
        qs = qs.filter(period_end__gte=date_from)
    if date_to:
        qs = qs.filter(period_start__lte=date_to)
    return qs


def get_feedback_queryset(date_from=None, date_to=None, unit=None):
    """Return RequestFeedback queryset for Director/GSO reports. Optional date range and unit."""
    from apps.gso_requests.models import RequestFeedback
    qs = RequestFeedback.objects.select_related(
        'request', 'user', 'request__unit', 'request__requestor',
    ).order_by('-created_at')
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
    if unit:
        qs = qs.filter(request__unit=unit)
    return qs


def build_feedback_export_excel(queryset):
    """Build Feedback (CSM) export Excel. Columns: Request ID, Unit, Requestor, Date, CC1-3, SQD1-9, Suggestions, Email."""
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = "Feedback"[:31]
    headers = [
        "Request ID", "Unit", "Requestor", "Submitted",
        "CC1", "CC2", "CC3",
        "SQD1", "SQD2", "SQD3", "SQD4", "SQD5", "SQD6", "SQD7", "SQD8", "SQD9",
        "Suggestions", "Email",
    ]
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    table_header_row = _apply_standard_header(
        ws,
        total_columns=len(headers),
        period_label="CUSTOMER SATISFACTION FEEDBACK",
        report_label="FEEDBACK EXPORT",
        unit_label="ALL UNITS",
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=table_header_row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        cell.fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    row = table_header_row + 1
    from django.utils import timezone
    for fb in queryset:
        req = fb.request
        cc1_display = dict(fb.CC1_CHOICES).get(fb.cc1, fb.cc1 or "")
        cc2_display = dict(fb.CC2_CHOICES).get(fb.cc2, fb.cc2 or "")
        cc3_display = dict(fb.CC3_CHOICES).get(fb.cc3, fb.cc3 or "")
        submitted_at = fb.created_at
        if timezone.is_aware(submitted_at):
            submitted_at = timezone.make_naive(submitted_at, timezone.get_current_timezone())
        ws.cell(row=row, column=1, value=req.display_id)
        ws.cell(row=row, column=2, value=req.unit.name if req.unit_id else "")
        ws.cell(row=row, column=3, value=fb.user.get_full_name() or fb.user.username)
        ws.cell(row=row, column=4, value=submitted_at)
        ws.cell(row=row, column=5, value=cc1_display)
        ws.cell(row=row, column=6, value=cc2_display)
        ws.cell(row=row, column=7, value=cc3_display)
        for i in range(1, 10):
            ws.cell(row=row, column=7 + i, value=getattr(fb, f'sqd{i}', None))
        ws.cell(row=row, column=17, value=fb.suggestions or "")
        ws.cell(row=row, column=18, value=fb.email or "")
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = thin_border
        row += 1
    feedback_widths = [
        14, 18, 20, 22,  # Request ID, Unit, Requestor, Submitted
        14, 14, 14,      # CC1-CC3
        8, 8, 8, 8, 8, 8, 8, 8, 8,  # SQD1-SQD9
        32, 24,          # Suggestions, Email
    ]
    for col in range(1, len(headers) + 1):
        width = feedback_widths[col - 1] if col - 1 < len(feedback_widths) else max(10, min(40, len(headers[col - 1]) + 2))
        ws.column_dimensions[get_column_letter(col)].width = width
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, "feedback"