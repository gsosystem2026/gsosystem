from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from openpyxl import load_workbook

from apps.gso_requests.models import Request
from apps.gso_units.models import Unit


@dataclass
class ImportStats:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: int = 0


def _norm(value) -> str:
    return str(value or "").strip()


def _norm_upper(value) -> str:
    return _norm(value).upper()


def _to_bool(value, default=False) -> bool:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    text = _norm_upper(value)
    if text in {"1", "TRUE", "YES", "Y", "ON"}:
        return True
    if text in {"0", "FALSE", "NO", "N", "OFF"}:
        return False
    return default


def _to_dt(value):
    if not value:
        return None
    if hasattr(value, "tzinfo"):
        dt = value
        if timezone.is_naive(dt):
            dt = timezone.make_aware(dt, timezone.get_current_timezone())
        return dt
    text = _norm(value)
    if not text:
        return None
    dt = timezone.datetime.fromisoformat(text)
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())
    return dt


class Command(BaseCommand):
    help = (
        "Import core migration data from Excel workbook. "
        "Supported sheets: Units, Users, Requests."
    )

    def add_arguments(self, parser):
        parser.add_argument("excel_file", help="Path to .xlsx migration file.")
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Validate and preview actions without writing changes.",
        )

    def handle(self, *args, **options):
        excel_path = Path(options["excel_file"]).expanduser().resolve()
        dry_run = bool(options["dry_run"])
        if not excel_path.exists():
            raise CommandError(f"File not found: {excel_path}")
        if excel_path.suffix.lower() != ".xlsx":
            raise CommandError("Only .xlsx files are supported.")

        wb = load_workbook(filename=str(excel_path), data_only=True)
        try:
            self.stdout.write(self.style.NOTICE(f"Reading workbook: {excel_path}"))
            self.stdout.write(self.style.NOTICE(f"Mode: {'DRY-RUN' if dry_run else 'APPLY'}"))

            unit_codes_in_file = set()
            usernames_in_file = set()
            stats = {
                "units": ImportStats(),
                "users": ImportStats(),
                "requests": ImportStats(),
            }

            with transaction.atomic():
                self._import_units(wb, stats["units"], dry_run, unit_codes_in_file)
                self._import_users(wb, stats["users"], dry_run, unit_codes_in_file, usernames_in_file)
                self._import_requests(wb, stats["requests"], dry_run, unit_codes_in_file, usernames_in_file)
                if dry_run:
                    transaction.set_rollback(True)

            self.stdout.write("")
            self.stdout.write(self.style.SUCCESS("Import summary"))
            for name, data in stats.items():
                self.stdout.write(
                    f"  {name}: created={data.created}, updated={data.updated}, "
                    f"skipped={data.skipped}, errors={data.errors}"
                )
            if any(s.errors for s in stats.values()):
                raise CommandError("Import completed with errors. Fix the workbook and retry.")
        finally:
            wb.close()

    def _sheet_rows(self, wb, sheet_name):
        if sheet_name not in wb.sheetnames:
            self.stdout.write(self.style.WARNING(f"Sheet '{sheet_name}' not found. Skipping."))
            return []
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [_norm(h).lower() for h in rows[0]]
        data_rows = []
        for idx, row in enumerate(rows[1:], start=2):
            payload = {headers[col]: row[col] for col in range(len(headers)) if headers[col]}
            payload["_row"] = idx
            data_rows.append(payload)
        return data_rows

    def _import_units(self, wb, stats: ImportStats, dry_run: bool, unit_codes_in_file: set[str]):
        rows = self._sheet_rows(wb, "Units")
        for row in rows:
            code = _norm(row.get("code")).lower()
            name = _norm(row.get("name"))
            if not code or not name:
                stats.errors += 1
                self.stdout.write(
                    self.style.ERROR(f"Units row {row['_row']}: 'code' and 'name' are required.")
                )
                continue
            unit_codes_in_file.add(code)
            defaults = {"name": name, "is_active": _to_bool(row.get("is_active"), True)}
            existing = Unit.objects.filter(code=code).first()
            if existing:
                changed = (
                    existing.name != defaults["name"]
                    or existing.is_active != defaults["is_active"]
                )
                if changed:
                    stats.updated += 1
                    if not dry_run:
                        Unit.objects.filter(pk=existing.pk).update(**defaults)
                else:
                    stats.skipped += 1
                continue
            stats.created += 1
            if not dry_run:
                Unit.objects.create(code=code, **defaults)

    def _import_users(
        self,
        wb,
        stats: ImportStats,
        dry_run: bool,
        unit_codes_in_file: set[str],
        usernames_in_file: set[str],
    ):
        User = get_user_model()
        rows = self._sheet_rows(wb, "Users")
        valid_roles = {r[0] for r in User.Role.choices}
        for row in rows:
            username = _norm(row.get("username"))
            email = _norm(row.get("email"))
            role = _norm_upper(row.get("role")) or User.Role.REQUESTOR
            if not username or not email:
                stats.errors += 1
                self.stdout.write(
                    self.style.ERROR(f"Users row {row['_row']}: 'username' and 'email' are required.")
                )
                continue
            if role not in valid_roles:
                stats.errors += 1
                self.stdout.write(
                    self.style.ERROR(f"Users row {row['_row']}: invalid role '{role}'.")
                )
                continue
            unit = None
            unit_code = _norm(row.get("unit_code")).lower()
            if unit_code:
                unit = Unit.objects.filter(code=unit_code).first()
                if not unit and unit_code not in unit_codes_in_file:
                    stats.errors += 1
                    self.stdout.write(
                        self.style.ERROR(
                            f"Users row {row['_row']}: unit_code '{unit_code}' does not exist."
                        )
                    )
                    continue
            payload = {
                "email": email,
                "first_name": _norm(row.get("first_name")),
                "last_name": _norm(row.get("last_name")),
                "role": role,
                "office_department": _norm(row.get("office_department")),
                "employment_status": _norm(row.get("employment_status")),
                "position_title": _norm(row.get("position_title")),
                "is_active": _to_bool(row.get("is_active"), True),
                "account_status": _norm_upper(row.get("account_status")) or "ACTIVE",
                "unit": unit,
            }

            existing = User.objects.filter(username__iexact=username).first()
            usernames_in_file.add(username.lower())
            if existing:
                changed = False
                for key, value in payload.items():
                    if getattr(existing, key) != value:
                        changed = True
                        break
                if changed:
                    stats.updated += 1
                    if not dry_run:
                        for key, value in payload.items():
                            setattr(existing, key, value)
                        existing.save()
                else:
                    stats.skipped += 1
                continue

            stats.created += 1
            if not dry_run:
                user = User(username=username, **payload)
                user.set_unusable_password()
                user.save()

    def _import_requests(
        self,
        wb,
        stats: ImportStats,
        dry_run: bool,
        unit_codes_in_file: set[str],
        usernames_in_file: set[str],
    ):
        rows = self._sheet_rows(wb, "Requests")
        valid_status = {s[0] for s in Request.Status.choices}
        User = get_user_model()
        for row in rows:
            title = _norm(row.get("title"))
            requestor_username = _norm(row.get("requestor_username"))
            unit_code = _norm(row.get("unit_code")).lower()
            status = _norm_upper(row.get("status")) or Request.Status.SUBMITTED
            if not title or not requestor_username or not unit_code:
                stats.errors += 1
                self.stdout.write(
                    self.style.ERROR(
                        f"Requests row {row['_row']}: 'title', 'requestor_username', and 'unit_code' are required."
                    )
                )
                continue
            if status not in valid_status:
                stats.errors += 1
                self.stdout.write(
                    self.style.ERROR(f"Requests row {row['_row']}: invalid status '{status}'.")
                )
                continue
            requestor = User.objects.filter(username__iexact=requestor_username).first()
            if not requestor and requestor_username.lower() not in usernames_in_file:
                stats.errors += 1
                self.stdout.write(
                    self.style.ERROR(
                        f"Requests row {row['_row']}: requestor '{requestor_username}' not found."
                    )
                )
                continue
            unit = Unit.objects.filter(code=unit_code).first()
            if not unit and unit_code not in unit_codes_in_file:
                stats.errors += 1
                self.stdout.write(
                    self.style.ERROR(
                        f"Requests row {row['_row']}: unit '{unit_code}' not found."
                    )
                )
                continue

            payload = {
                "title": title,
                "description": _norm(row.get("description")),
                "location": _norm(row.get("location")),
                "status": status,
                "labor": _to_bool(row.get("labor"), False),
                "materials": _to_bool(row.get("materials"), False),
                "others": _to_bool(row.get("others"), False),
                "is_emergency": _to_bool(row.get("is_emergency"), False),
            }
            request_id_raw = _norm(row.get("request_id"))
            existing = None
            if request_id_raw.isdigit():
                existing = Request.objects.filter(pk=int(request_id_raw)).first()

            if existing:
                changed = False
                if requestor and existing.requestor_id != requestor.id:
                    changed = True
                if unit and existing.unit_id != unit.id:
                    changed = True
                for key, value in payload.items():
                    if getattr(existing, key) != value:
                        changed = True
                        break
                if changed:
                    stats.updated += 1
                    if not dry_run:
                        if requestor:
                            existing.requestor = requestor
                        if unit:
                            existing.unit = unit
                        for key, value in payload.items():
                            setattr(existing, key, value)
                        existing.save()
                else:
                    stats.skipped += 1
                continue

            stats.created += 1
            if not dry_run:
                if not requestor or not unit:
                    # In apply mode these must exist in DB.
                    stats.errors += 1
                    self.stdout.write(
                        self.style.ERROR(
                            f"Requests row {row['_row']}: dependencies missing in DB for create."
                        )
                    )
                    continue
                req = Request.objects.create(
                    requestor=requestor,
                    unit=unit,
                    **payload,
                )
                created_at = _to_dt(row.get("created_at"))
                updated_at = _to_dt(row.get("updated_at"))
                updates = {}
                if created_at:
                    updates["created_at"] = created_at
                if updated_at:
                    updates["updated_at"] = updated_at
                if updates:
                    Request.objects.filter(pk=req.pk).update(**updates)
