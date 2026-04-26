from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from apps.gso_reports.models import IPMTDraft
from apps.gso_units.models import Unit


@dataclass
class Stats:
    drafts_created: int = 0
    drafts_updated: int = 0
    rows_parsed: int = 0
    rows_skipped: int = 0
    errors: int = 0


def _norm(value) -> str:
    return str(value or "").strip()


class Command(BaseCommand):
    help = (
        "Import legacy IPMT workbook template into IPMTDraft. "
        "Reads unit/employee/month metadata and indicator rows."
    )

    def add_arguments(self, parser):
        parser.add_argument("excel_file", help="Path to legacy IPMT .xlsx file.")
        parser.add_argument("--unit-code", required=True, help="Target unit code, e.g. repair")
        parser.add_argument(
            "--updated-by-username",
            default="migrated_requestor",
            help="Username to set as IPMT draft updater when available.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Validate and preview without writing changes.",
        )

    def handle(self, *args, **options):
        excel_path = Path(options["excel_file"]).expanduser().resolve()
        if not excel_path.exists():
            raise CommandError(f"File not found: {excel_path}")
        if excel_path.suffix.lower() != ".xlsx":
            raise CommandError("Only .xlsx files are supported.")

        dry_run = bool(options["dry_run"])
        unit_code = _norm(options["unit_code"]).lower()
        unit = Unit.objects.filter(code=unit_code, is_active=True).first()
        if not unit:
            raise CommandError(f"Unit not found for code '{unit_code}'.")

        wb = load_workbook(filename=str(excel_path), data_only=True)
        stats = Stats()
        try:
            ws = wb.worksheets[0]
            metadata = self._parse_metadata(ws)
            detected_unit = self._match_unit_from_text(metadata["unit_text"])
            if detected_unit and detected_unit.id != unit.id:
                raise CommandError(
                    f"Workbook appears to be for unit '{detected_unit.code}' ({detected_unit.name}) "
                    f"but selected target unit is '{unit.code}' ({unit.name})."
                )
            personnel = self._resolve_personnel(metadata["employee_name"], unit)
            if not personnel:
                raise CommandError(
                    f"Could not map employee '{metadata['employee_name']}' to active personnel in unit '{unit.code}'."
                )
            year, month = self._parse_month_year(metadata["month_text"])
            rows_json = self._parse_rows(ws, stats)
            if not rows_json:
                raise CommandError("No valid IPMT rows found under the indicator table.")

            updater = self._resolve_updater(options["updated_by_username"])

            self.stdout.write(self.style.NOTICE(f"Workbook: {excel_path}"))
            self.stdout.write(self.style.NOTICE(f"Mode: {'DRY-RUN' if dry_run else 'APPLY'}"))
            self.stdout.write(self.style.NOTICE(f"Target unit: {unit.code} ({unit.name})"))
            self.stdout.write(self.style.NOTICE(f"Employee mapped: {personnel.get_full_name() or personnel.username}"))
            self.stdout.write(self.style.NOTICE(f"Period mapped: {year}-{month:02d}"))

            with transaction.atomic():
                if dry_run:
                    stats.drafts_created = 1 if not IPMTDraft.objects.filter(
                        personnel=personnel, year=year, month=month
                    ).exists() else 0
                    stats.drafts_updated = 0 if stats.drafts_created else 1
                    transaction.set_rollback(True)
                else:
                    _, created = IPMTDraft.objects.update_or_create(
                        personnel=personnel,
                        year=year,
                        month=month,
                        defaults={
                            "rows_json": rows_json,
                            "updated_by": updater,
                        },
                    )
                    if created:
                        stats.drafts_created += 1
                    else:
                        stats.drafts_updated += 1
        finally:
            wb.close()

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("Legacy IPMT import summary"))
        self.stdout.write(f"  Drafts created: {stats.drafts_created}")
        self.stdout.write(f"  Drafts updated: {stats.drafts_updated}")
        self.stdout.write(f"  Rows parsed: {stats.rows_parsed}")
        self.stdout.write(f"  Rows skipped: {stats.rows_skipped}")
        self.stdout.write(f"  Errors: {stats.errors}")
        if stats.errors:
            raise CommandError("Import finished with errors. Fix data and retry.")

    def _parse_metadata(self, ws):
        unit_text = _norm(ws.cell(row=5, column=2).value)
        employee_name = _norm(ws.cell(row=6, column=2).value)
        month_text = _norm(ws.cell(row=9, column=2).value)
        if not unit_text:
            raise CommandError("Missing unit value in row 5.")
        if not employee_name:
            raise CommandError("Missing employee name in row 6.")
        if not month_text:
            raise CommandError("Missing month value in row 9.")
        return {
            "unit_text": unit_text,
            "employee_name": employee_name,
            "month_text": month_text,
        }

    def _match_unit_from_text(self, text):
        unit_text = _norm(text).lower()
        if not unit_text:
            return None
        units = Unit.objects.filter(is_active=True)
        for unit in units:
            if _norm(unit.name).lower() in unit_text or _norm(unit.code).lower() in unit_text:
                return unit
        return None

    def _resolve_personnel(self, employee_name, unit):
        User = get_user_model()
        name_norm = _norm(employee_name).lower()
        qs = User.objects.filter(role=User.Role.PERSONNEL, is_active=True, unit=unit)
        for user in qs:
            full_name = f"{_norm(user.first_name)} {_norm(user.last_name)}".strip().lower()
            if full_name and full_name == name_norm:
                return user
            if _norm(user.username).lower() == name_norm:
                return user
        return None

    def _parse_month_year(self, month_text):
        text = _norm(month_text).upper()
        # Fallback: detect month token and 4-digit year
        month_num = None
        for idx, name in enumerate(
            [
                "JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE",
                "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER",
            ],
            start=1,
        ):
            if name in text:
                month_num = idx
                break
        year_num = None
        for token in text.replace(",", " ").split():
            if token.isdigit() and len(token) == 4:
                year_num = int(token)
                break
        if month_num and year_num:
            return year_num, month_num
        raise CommandError(f"Unable to parse month/year from row 9 value: '{month_text}'")

    def _parse_rows(self, ws, stats: Stats):
        header_row = None
        for idx in range(1, min(30, ws.max_row) + 1):
            left = _norm(ws.cell(row=idx, column=1).value).lower()
            mid = _norm(ws.cell(row=idx, column=2).value).lower()
            if "success indicators" in left and "actual accomplishments" in mid:
                header_row = idx
                break
        if not header_row:
            raise CommandError("IPMT table header not found.")

        grouped = []
        current_indicator = ""
        for row_idx in range(header_row + 1, ws.max_row + 1):
            indicator = _norm(ws.cell(row=row_idx, column=1).value)
            accomplishment = _norm(ws.cell(row=row_idx, column=2).value)
            comment = _norm(ws.cell(row=row_idx, column=3).value) or "Complied"
            if indicator.startswith("*Based on the IPCR"):
                break
            if indicator:
                current_indicator = indicator
            if not accomplishment:
                if indicator:
                    stats.rows_skipped += 1
                continue
            if not current_indicator:
                stats.rows_skipped += 1
                continue
            stats.rows_parsed += 1
            existing = next((x for x in grouped if x["indicator"] == current_indicator), None)
            if not existing:
                grouped.append(
                    {
                        "indicator": current_indicator,
                        "accomplishments": [accomplishment],
                        "comment": comment,
                    }
                )
            else:
                existing["accomplishments"].append(accomplishment)
                if not existing.get("comment") and comment:
                    existing["comment"] = comment
        return grouped

    def _resolve_updater(self, username):
        username = _norm(username).lower()
        if not username:
            return None
        User = get_user_model()
        return User.objects.filter(username__iexact=username).first()
