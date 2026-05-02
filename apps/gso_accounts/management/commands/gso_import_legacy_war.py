from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import re

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from openpyxl import load_workbook

from apps.gso_accounts.legacy_migration_workbook import (
    find_war_header_row,
    workbook_has_ipmt_fingerprint,
    workbook_has_war_header,
)
from apps.gso_requests.models import Request, RequestAssignment
from apps.gso_reports.models import WorkAccomplishmentReport
from apps.gso_units.models import Unit


@dataclass
class Stats:
    requests_created: int = 0
    requests_skipped: int = 0
    wars_created: int = 0
    wars_skipped: int = 0
    requestor_mapped_existing: int = 0
    requestor_created_from_office: int = 0
    requestor_fallback_blank_office: int = 0
    personnel_mapped_existing: int = 0
    personnel_created_from_name: int = 0
    personnel_fallback_blank_name: int = 0
    errors: int = 0


def _norm(value) -> str:
    return str(value or "").strip()


def _parse_date(value):
    if value is None or value == "":
        return None
    if hasattr(value, "date"):
        return value.date()
    text = _norm(value)
    if not text:
        return None
    text = text.replace("//", "/")
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal(value):
    text = _norm(value)
    if not text:
        return None
    cleaned = text.replace(",", "")
    if cleaned.upper() in {"NO JO", "DIFF FORM"}:
        return None
    try:
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None


def _office_key(value) -> str:
    text = _norm(value).lower()
    return "".join(ch for ch in text if ch.isalnum())


def _username_slug(value) -> str:
    return re.sub(r"[^a-z0-9]+", "_", _norm(value).lower()).strip("_")


def _personnel_key(value) -> str:
    return _office_key(value)


class Command(BaseCommand):
    help = (
        "Import legacy WAR workbook (monthly sheets) using placeholder migrated users. "
        "Creates completed request stubs + WAR rows."
    )

    def add_arguments(self, parser):
        parser.add_argument("excel_file", help="Path to legacy WAR .xlsx file.")
        parser.add_argument("--unit-code", required=True, help="Target unit code, e.g. electrical")
        parser.add_argument(
            "--legacy-personnel-username",
            default="migrated_legacy",
            help="Placeholder personnel username for unmapped records.",
        )
        parser.add_argument(
            "--legacy-requestor-username",
            default="migrated_requestor",
            help="Placeholder requestor username for created request stubs.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Validate and preview without writing changes.",
        )

    def handle(self, *args, **options):
        excel_path = Path(options["excel_file"]).expanduser().resolve()
        if not excel_path.exists():
            raise CommandError(f"File not found: {excel_path}")
        if excel_path.suffix.lower() != ".xlsx":
            raise CommandError("Only .xlsx files are supported.")

        dry_run = bool(options["dry_run"])
        unit_code = _norm(options["unit_code"]).lower()
        if not unit_code:
            raise CommandError("--unit-code is required.")
        unit = Unit.objects.filter(code=unit_code).first()
        if not unit:
            raise CommandError(f"Unit not found for code '{unit_code}'.")

        wb = load_workbook(filename=str(excel_path), data_only=True)
        stats = Stats()
        try:
            has_ipmt = workbook_has_ipmt_fingerprint(wb)
            has_war = workbook_has_war_header(wb)
            if has_ipmt and has_war:
                raise CommandError(
                    "This workbook looks like multiple templates at once "
                    "(IPMT layout and a WAR data table). Use a separate file per report type."
                )
            if has_ipmt and not has_war:
                raise CommandError(
                    'This workbook matches the legacy IPMT Excel template. '
                    'Choose Report Type "IPMT", not WAR.'
                )
            if not has_war:
                raise CommandError(
                    "No legacy Work Accomplishment Report layout was found "
                    '(expected columns such as Date Started, Description, Requesting Office, '
                    '"Name of Activity" or "Name of Project", Date Completed).'
                )

            detected_unit = self._detect_unit_from_workbook(wb)
            if detected_unit and detected_unit.id != unit.id:
                raise CommandError(
                    f"Workbook appears to be for unit '{detected_unit.code}' ({detected_unit.name}) "
                    f"but selected target unit is '{unit.code}' ({unit.name})."
                )

            User = get_user_model()
            legacy_personnel = self._ensure_legacy_personnel(
                User, options["legacy_personnel_username"], unit, dry_run
            )
            legacy_requestor = self._ensure_legacy_requestor(
                User, options["legacy_requestor_username"], dry_run
            )
            requestor_office_map = self._build_requestor_office_map(User)
            personnel_name_map = self._build_personnel_name_map(User, unit)

            self.stdout.write(self.style.NOTICE(f"Workbook: {excel_path}"))
            self.stdout.write(self.style.NOTICE(f"Mode: {'DRY-RUN' if dry_run else 'APPLY'}"))
            self.stdout.write(self.style.NOTICE(f"Target unit: {unit.code} ({unit.name})"))
            if detected_unit:
                self.stdout.write(
                    self.style.NOTICE(
                        f"Detected workbook unit: {detected_unit.code} ({detected_unit.name})"
                    )
                )
            self.stdout.write(self.style.NOTICE(f"Legacy personnel: {legacy_personnel.username}"))
            self.stdout.write(self.style.NOTICE(f"Legacy requestor: {legacy_requestor.username}"))

            with transaction.atomic():
                for sheet in wb.worksheets:
                    self._import_sheet(
                        sheet=sheet,
                        workbook_name=excel_path.name,
                        unit=unit,
                        legacy_personnel=legacy_personnel,
                        legacy_requestor=legacy_requestor,
                        requestor_office_map=requestor_office_map,
                        personnel_name_map=personnel_name_map,
                        dry_run=dry_run,
                        stats=stats,
                    )
                if dry_run:
                    transaction.set_rollback(True)
        finally:
            wb.close()

        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("Legacy WAR import summary"))
        self.stdout.write(f"  Requests created: {stats.requests_created}")
        self.stdout.write(f"  Requests skipped: {stats.requests_skipped}")
        self.stdout.write(f"  WAR created: {stats.wars_created}")
        self.stdout.write(f"  WAR skipped: {stats.wars_skipped}")
        self.stdout.write(f"  Requestor mapped (existing user): {stats.requestor_mapped_existing}")
        self.stdout.write(f"  Requestor created from office text: {stats.requestor_created_from_office}")
        self.stdout.write(f"  Requestor fallback used (blank office): {stats.requestor_fallback_blank_office}")
        self.stdout.write(f"  Personnel mapped (existing user): {stats.personnel_mapped_existing}")
        self.stdout.write(f"  Personnel created from name text: {stats.personnel_created_from_name}")
        self.stdout.write(f"  Personnel fallback used (blank name): {stats.personnel_fallback_blank_name}")
        self.stdout.write(f"  Errors: {stats.errors}")
        if stats.errors:
            raise CommandError("Import finished with errors. Fix data and retry.")

    def _detect_unit_from_workbook(self, wb):
        units = list(Unit.objects.filter(is_active=True))
        if not units:
            return None
        score_by_unit_id = {u.id: 0 for u in units}
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=min(12, sheet.max_row), values_only=True):
                for cell in row:
                    text = _norm(cell).lower()
                    if not text:
                        continue
                    for unit in units:
                        code = _norm(unit.code).lower()
                        name = _norm(unit.name).lower()
                        if code and re.search(rf"\b{re.escape(code)}\b", text):
                            score_by_unit_id[unit.id] += 2
                        if name and re.search(rf"\b{re.escape(name)}\b", text):
                            score_by_unit_id[unit.id] += 3
        max_score = max(score_by_unit_id.values(), default=0)
        if max_score <= 0:
            return None
        winners = [u for u in units if score_by_unit_id.get(u.id) == max_score]
        if len(winners) != 1:
            return None
        return winners[0]

    def _ensure_legacy_personnel(self, User, username, unit, dry_run):
        username = _norm(username).lower()
        user = User.objects.filter(username__iexact=username).first()
        if user:
            if user.role != User.Role.PERSONNEL:
                raise CommandError(
                    f"User '{username}' exists but role is {user.role}, expected PERSONNEL."
                )
            if user.unit_id != unit.id and not dry_run:
                user.unit = unit
                user.save(update_fields=["unit"])
            return user
        if dry_run:
            fake = User(username=username, role=User.Role.PERSONNEL, unit=unit)
            fake.first_name = "Migrated"
            fake.last_name = "Legacy"
            return fake
        user = User.objects.create_user(
            username=username,
            email=f"{username}@legacy.local",
            first_name="Migrated",
            last_name="Legacy",
            role=User.Role.PERSONNEL,
            unit=unit,
            is_active=True,
            employment_status="Legacy Import",
            position_title="Unmapped Legacy Personnel",
        )
        user.set_unusable_password()
        user.save(update_fields=["password"])
        return user

    def _ensure_legacy_requestor(self, User, username, dry_run):
        username = _norm(username).lower()
        user = User.objects.filter(username__iexact=username).first()
        if user:
            if user.role != User.Role.REQUESTOR:
                raise CommandError(
                    f"User '{username}' exists but role is {user.role}, expected REQUESTOR."
                )
            return user
        if dry_run:
            fake = User(username=username, role=User.Role.REQUESTOR)
            fake.first_name = "Migrated"
            fake.last_name = "Requestor"
            return fake
        user = User.objects.create_user(
            username=username,
            email=f"{username}@legacy.local",
            first_name="Migrated",
            last_name="Requestor",
            role=User.Role.REQUESTOR,
            is_active=True,
            office_department="Migrated Legacy Records",
        )
        user.set_unusable_password()
        user.save(update_fields=["password"])
        return user

    def _find_header_row(self, sheet):
        return find_war_header_row(sheet)

    def _build_requestor_office_map(self, User):
        office_map = {}
        qs = User.objects.filter(role=User.Role.REQUESTOR).exclude(office_department="")
        for user in qs:
            key = _office_key(user.office_department)
            if not key:
                continue
            # Keep first mapping to avoid ambiguous override.
            office_map.setdefault(key, user)
        return office_map

    def _build_personnel_name_map(self, User, unit):
        name_map = {}
        qs = User.objects.filter(role=User.Role.PERSONNEL, unit=unit)
        for user in qs:
            full_name = f"{_norm(user.first_name)} {_norm(user.last_name)}".strip()
            for token in (full_name, user.username):
                key = _personnel_key(token)
                if key:
                    name_map.setdefault(key, user)
        return name_map

    def _resolve_requestor(self, requesting_office, office_map, legacy_requestor):
        key = _office_key(requesting_office)
        if key and key in office_map:
            return office_map[key], "mapped_existing"
        if not key:
            return legacy_requestor, "fallback_blank"
        return None, "create_from_office"

    def _ensure_requestor_for_office(self, User, requesting_office, dry_run):
        key_slug = _username_slug(requesting_office) or "office"
        username = f"migrated_req_{key_slug}"[:150]
        user = User.objects.filter(username__iexact=username).first()
        if user:
            return user
        if dry_run:
            fake = User(username=username, role=User.Role.REQUESTOR)
            fake.first_name = "Migrated"
            fake.last_name = "Office"
            fake.office_department = _norm(requesting_office)[:255]
            return fake
        user = User.objects.create_user(
            username=username,
            email=f"{username}@legacy.local",
            first_name="Migrated",
            last_name="Office",
            role=User.Role.REQUESTOR,
            is_active=True,
            office_department=_norm(requesting_office)[:255] or "Migrated Legacy Records",
        )
        user.set_unusable_password()
        user.save(update_fields=["password"])
        return user

    def _resolve_personnel(self, assigned_personnel, personnel_map, legacy_personnel):
        key = _personnel_key(assigned_personnel)
        if key and key in personnel_map:
            return personnel_map[key], "mapped_existing"
        if not key:
            return legacy_personnel, "fallback_blank"
        return None, "create_from_name"

    def _ensure_personnel_for_name(self, User, assigned_personnel, unit, dry_run):
        key_slug = _username_slug(assigned_personnel) or "personnel"
        username = f"migrated_per_{key_slug}"[:150]
        user = User.objects.filter(username__iexact=username).first()
        if user:
            if user.role != User.Role.PERSONNEL:
                raise CommandError(
                    f"User '{username}' exists but role is {user.role}, expected PERSONNEL."
                )
            return user
        first_name = _norm(assigned_personnel)
        if dry_run:
            fake = User(username=username, role=User.Role.PERSONNEL, unit=unit)
            fake.first_name = first_name[:150] or "Migrated"
            fake.last_name = "Legacy"
            return fake
        user = User.objects.create_user(
            username=username,
            email=f"{username}@legacy.local",
            first_name=first_name[:150] or "Migrated",
            last_name="Legacy",
            role=User.Role.PERSONNEL,
            unit=unit,
            is_active=True,
            employment_status="Legacy Import",
            position_title="Imported from legacy WAR assigned personnel",
        )
        user.set_unusable_password()
        user.save(update_fields=["password"])
        return user

    def _import_sheet(
        self,
        sheet,
        workbook_name,
        unit,
        legacy_personnel,
        legacy_requestor,
        requestor_office_map,
        personnel_name_map,
        dry_run,
        stats: Stats,
    ):
        header_row = self._find_header_row(sheet)
        if not header_row:
            self.stdout.write(self.style.WARNING(f"Skipping sheet '{sheet.title}': header not found."))
            return

        self.stdout.write(self.style.NOTICE(f"Processing sheet '{sheet.title}' (header row {header_row})"))
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            row = [sheet.cell(row=row_idx, column=c).value for c in range(1, 12)]
            if all(_norm(v) == "" for v in row):
                continue
            first_col = _norm(row[0]).lower()
            if first_col == "total":
                continue

            date_started = _parse_date(row[0])
            date_completed = _parse_date(row[1]) or date_started
            activity = _norm(row[2])
            description = _norm(row[3])
            requesting_office = _norm(row[4])
            assigned_personnel = _norm(row[5])
            status_raw = _norm(row[6]) or "Done"
            material_cost = _parse_decimal(row[7])
            labor_cost = _parse_decimal(row[8])
            total_cost = _parse_decimal(row[9])
            legacy_control = _norm(row[10])

            if not activity and not description:
                stats.wars_skipped += 1
                continue
            if not date_started:
                stats.wars_skipped += 1
                self.stdout.write(
                    self.style.WARNING(
                        f"  Row {row_idx}: missing/invalid Date Started, skipped as non-importable legacy row."
                    )
                )
                continue

            marker = (
                f"[MIGRATED-LEGACY-WAR|FILE:{workbook_name}|SHEET:{sheet.title}|ROW:{row_idx}]"
            )
            existing_request = Request.objects.filter(description__icontains=marker).first()
            selected_personnel = legacy_personnel
            if existing_request:
                stats.requests_skipped += 1
                request_obj = existing_request
            else:
                stats.requests_created += 1
                selected_requestor, source = self._resolve_requestor(
                    requesting_office=requesting_office,
                    office_map=requestor_office_map,
                    legacy_requestor=legacy_requestor,
                )
                if source == "mapped_existing":
                    stats.requestor_mapped_existing += 1
                elif source == "create_from_office":
                    selected_requestor = self._ensure_requestor_for_office(
                        User=get_user_model(),
                        requesting_office=requesting_office,
                        dry_run=dry_run,
                    )
                    requestor_office_map[_office_key(requesting_office)] = selected_requestor
                    stats.requestor_created_from_office += 1
                else:
                    stats.requestor_fallback_blank_office += 1

                selected_personnel, personnel_source = self._resolve_personnel(
                    assigned_personnel=assigned_personnel,
                    personnel_map=personnel_name_map,
                    legacy_personnel=legacy_personnel,
                )
                if personnel_source == "mapped_existing":
                    stats.personnel_mapped_existing += 1
                elif personnel_source == "create_from_name":
                    selected_personnel = self._ensure_personnel_for_name(
                        User=get_user_model(),
                        assigned_personnel=assigned_personnel,
                        unit=unit,
                        dry_run=dry_run,
                    )
                    personnel_name_map[_personnel_key(assigned_personnel)] = selected_personnel
                    stats.personnel_created_from_name += 1
                else:
                    stats.personnel_fallback_blank_name += 1

                if dry_run:
                    request_obj = None
                else:
                    request_description = (
                        f"{marker}\n"
                        f"Legacy Control #: {legacy_control or 'N/A'}\n"
                        f"Legacy Office: {requesting_office or 'N/A'}\n"
                        f"Legacy Status: {status_raw or 'N/A'}\n\n"
                        f"{description}"
                    ).strip()
                    request_obj = Request.objects.create(
                        requestor=selected_requestor,
                        unit=unit,
                        title=activity[:255] or "Migrated legacy activity",
                        description=request_description,
                        location=requesting_office[:255],
                        status=Request.Status.COMPLETED,
                        labor=bool(material_cost or labor_cost or total_cost),
                        materials=bool(material_cost),
                        others=False,
                        is_emergency=False,
                    )
                    RequestAssignment.objects.get_or_create(
                        request=request_obj,
                        personnel=selected_personnel,
                        defaults={"assigned_by": None},
                    )
                    dt_start = timezone.make_aware(
                        datetime.combine(date_started, datetime.min.time()),
                        timezone.get_current_timezone(),
                    )
                    dt_end = timezone.make_aware(
                        datetime.combine(date_completed or date_started, datetime.min.time()),
                        timezone.get_current_timezone(),
                    )
                    Request.objects.filter(pk=request_obj.pk).update(
                        created_at=dt_start,
                        updated_at=dt_end,
                    )

            if request_obj is None:
                stats.wars_created += 1
                continue

            existing_war = WorkAccomplishmentReport.objects.filter(request=request_obj).first()
            if existing_war:
                stats.wars_skipped += 1
                continue

            stats.wars_created += 1
            if dry_run:
                continue

            accomplishments = (description or "").strip()
            assigned_personnel_obj = (
                RequestAssignment.objects.filter(request=request_obj)
                .select_related("personnel")
                .first()
            )
            WorkAccomplishmentReport.objects.create(
                request=request_obj,
                personnel=assigned_personnel_obj.personnel if assigned_personnel_obj else legacy_personnel,
                period_start=date_started,
                period_end=date_completed or date_started,
                summary=(activity or "Migrated legacy activity")[:255],
                accomplishments=accomplishments,
                material_cost=material_cost,
                labor_cost=labor_cost,
                total_cost=total_cost,
                created_by=legacy_requestor,
            )
