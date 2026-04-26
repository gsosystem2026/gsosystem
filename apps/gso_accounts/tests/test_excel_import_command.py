import os
import tempfile
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.test import TestCase

from openpyxl import Workbook

from apps.gso_requests.models import Request
from apps.gso_units.models import Unit


class ExcelImportCommandTests(TestCase):
    def _build_workbook(self):
        wb = Workbook()
        ws_units = wb.active
        ws_units.title = "Units"
        ws_units.append(["code", "name", "is_active"])
        ws_units.append(["repair", "Repair and Maintenance", True])

        ws_users = wb.create_sheet("Users")
        ws_users.append(
            [
                "username",
                "email",
                "first_name",
                "last_name",
                "role",
                "unit_code",
                "office_department",
                "employment_status",
                "position_title",
                "is_active",
                "account_status",
            ]
        )
        ws_users.append(
            [
                "req1",
                "req1@example.com",
                "Req",
                "One",
                "REQUESTOR",
                "",
                "Registrar",
                "",
                "",
                True,
                "ACTIVE",
            ]
        )

        ws_requests = wb.create_sheet("Requests")
        ws_requests.append(
            [
                "title",
                "description",
                "location",
                "requestor_username",
                "unit_code",
                "status",
                "labor",
                "materials",
                "others",
                "is_emergency",
            ]
        )
        ws_requests.append(
            [
                "Fix chair",
                "Please repair chair in room 2",
                "Room 2",
                "req1",
                "repair",
                "SUBMITTED",
                True,
                False,
                False,
                False,
            ]
        )
        return wb

    def test_dry_run_does_not_write(self):
        wb = self._build_workbook()
        fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        path = Path(temp_path)
        try:
            wb.save(path)
            wb.close()
            call_command("gso_import_excel", str(path), "--dry-run")
        finally:
            try:
                path.unlink(missing_ok=True)
            except PermissionError:
                pass
        self.assertEqual(Unit.objects.count(), 0)
        self.assertEqual(get_user_model().objects.count(), 0)
        self.assertEqual(Request.objects.count(), 0)

    def test_apply_creates_core_records(self):
        wb = self._build_workbook()
        fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        path = Path(temp_path)
        try:
            wb.save(path)
            wb.close()
            call_command("gso_import_excel", str(path))
        finally:
            try:
                path.unlink(missing_ok=True)
            except PermissionError:
                pass

        unit = Unit.objects.get(code="repair")
        user = get_user_model().objects.get(username="req1")
        req = Request.objects.get(title="Fix chair")
        self.assertEqual(req.unit_id, unit.id)
        self.assertEqual(req.requestor_id, user.id)
