import io
import tempfile
from pathlib import Path

from django.contrib.auth import get_user_model
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from apps.gso_requests.models import Request, MotorpoolTripData
from apps.gso_units.models import Unit


@override_settings(
    STORAGES={
        "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
        "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
    }
)
class MotorpoolRequestorPrintAccessTests(TestCase):
    """Owning requestors can open Motorpool HTML print routes; others get 404."""

    def setUp(self):
        self.motorpool = Unit.objects.create(name="Motorpool", code="motorpool", is_active=True)
        self.utility = Unit.objects.create(name="Utility", code="utility", is_active=True)
        User = get_user_model()
        self.owner = User.objects.create_user(
            username="mp_owner",
            password="Pass1234!",
            role=User.Role.REQUESTOR,
            email="mp-owner@example.com",
        )
        self.other_requestor = User.objects.create_user(
            username="mp_other",
            password="Pass1234!",
            role=User.Role.REQUESTOR,
            email="mp-other@example.com",
        )
        self.req = Request.objects.create(
            requestor=self.owner,
            unit=self.motorpool,
            title="Trip for meeting",
            description="Campus visit",
            status=Request.Status.SUBMITTED,
        )
        MotorpoolTripData.objects.create(request=self.req)

    def test_owner_can_print_request_and_trip_ticket(self):
        self.client.login(username="mp_owner", password="Pass1234!")
        for name in ("gso_requests:motorpool_print_request", "gso_requests:motorpool_print_trip_ticket"):
            res = self.client.get(reverse(name, kwargs={"pk": self.req.pk}))
            self.assertEqual(res.status_code, 200, msg=name)

    def test_non_owner_requestor_gets_404(self):
        self.client.login(username="mp_other", password="Pass1234!")
        url = reverse("gso_requests:motorpool_print_request", kwargs={"pk": self.req.pk})
        self.assertEqual(self.client.get(url).status_code, 404)

    def test_non_motorpool_request_404_even_for_owner(self):
        non_mp = Request.objects.create(
            requestor=self.owner,
            unit=self.utility,
            title="Leak fix",
            description="Washroom tap",
            location="Block A",
            status=Request.Status.SUBMITTED,
        )
        self.client.login(username="mp_owner", password="Pass1234!")
        url = reverse("gso_requests:motorpool_print_request", kwargs={"pk": non_mp.pk})
        self.assertEqual(self.client.get(url).status_code, 404)

    def test_owner_can_download_request_excel(self):
        tmp = Path(tempfile.gettempdir()) / "gso_test_request_for_vehicle.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "VEHICLE REQUEST"
        # Minimal cells required by mapper (copy 1 and copy 2 anchor rows).
        ws["C6"] = ""
        ws["J6"] = ""
        ws["D11"] = ""
        ws["D12"] = ""
        ws["D14"] = ""
        ws["D15"] = ""
        ws["D16"] = ""
        ws["E19"] = ""
        ws["B22"] = ""
        ws["C38"] = ""
        ws["J38"] = ""
        ws["D43"] = ""
        ws["D44"] = ""
        ws["D46"] = ""
        ws["D47"] = ""
        ws["D48"] = ""
        ws["E51"] = ""
        ws["B54"] = ""
        wb.save(tmp)

        mp = MotorpoolTripData.objects.get(request=self.req)
        mp.requesting_office = "Registrar"
        mp.places_to_be_visited = "City Hall"
        mp.itinerary_of_travel = "PSU -> City Hall"
        mp.number_of_days = 2
        mp.number_of_passengers = 5
        mp.save()

        self.client.login(username="mp_owner", password="Pass1234!")
        with self.settings(GSO_MOTORPOOL_REQUEST_XLSX_TEMPLATE=str(tmp)):
            url = reverse("gso_requests:motorpool_request_excel", kwargs={"pk": self.req.pk})
            res = self.client.get(url)
        self.assertEqual(res.status_code, 200)
        self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", res["Content-Type"])
        out_wb = load_workbook(io.BytesIO(res.content))
        out_ws = out_wb[out_wb.sheetnames[0]]
        # Filled in both copies.
        self.assertEqual(out_ws["D11"].value, "Registrar")
        self.assertEqual(out_ws["D43"].value, "Registrar")
        self.assertEqual(out_ws["D14"].value, "City Hall")
        self.assertEqual(out_ws["D46"].value, "City Hall")

    def test_owner_can_download_trip_ticket_excel(self):
        tmp = Path(tempfile.gettempdir()) / "gso_test_trip_ticket.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "TRIP TICKET"
        # Minimal anchor cells used by exporter mapping.
        ws["C9"] = ""
        ws["I9"] = ""
        ws["C10"] = ""
        ws["C11"] = ""
        ws["H11"] = ""
        ws["E12"] = ""
        ws["C14"] = ""
        ws["B22"] = ""
        ws["C22"] = ""
        ws["D22"] = ""
        ws["E22"] = ""
        ws["F22"] = ""
        ws["G22"] = ""
        ws["H22"] = ""
        ws["I22"] = ""
        ws["J22"] = ""
        ws["E47"] = ""
        ws["E48"] = ""
        ws["E49"] = ""
        ws["E50"] = ""
        ws["E51"] = ""
        ws["E52"] = ""
        ws["C54"] = ""
        wb.save(tmp)

        mp = MotorpoolTripData.objects.get(request=self.req)
        mp.requesting_office = "Registrar"
        mp.contact_number = "09123456789"
        mp.driver_name = "Juan Driver"
        mp.vehicle_plate = "ABC-123"
        mp.number_of_passengers = 3
        mp.fuel_beginning_liters = 10
        mp.other_consumables_notes = "None"
        mp.actual_legs_json = [
            {"depart_place": "PSU", "arrive_place": "City Hall", "distance": "5"}
        ]
        mp.save()

        self.client.login(username="mp_owner", password="Pass1234!")
        with self.settings(GSO_MOTORPOOL_TRIP_TICKET_XLSX_TEMPLATE=str(tmp)):
            url = reverse("gso_requests:motorpool_trip_ticket_excel", kwargs={"pk": self.req.pk})
            res = self.client.get(url)
        self.assertEqual(res.status_code, 200)
        self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", res["Content-Type"])
        out_wb = load_workbook(io.BytesIO(res.content))
        out_ws = out_wb[out_wb.sheetnames[0]]
        self.assertEqual(out_ws["C11"].value, "Juan Driver")
        self.assertEqual(out_ws["H11"].value, "ABC-123")
        self.assertEqual(out_ws["G22"].value, "Registrar")
        self.assertEqual(out_ws["H22"].value, "City Hall")

    def test_non_owner_cannot_download_trip_ticket_excel(self):
        self.client.login(username="mp_other", password="Pass1234!")
        url = reverse("gso_requests:motorpool_trip_ticket_excel", kwargs={"pk": self.req.pk})
        self.assertEqual(self.client.get(url).status_code, 404)
